import cv2
import numpy as np
import dlib
import face_recognition
import os
from tqdm import tqdm
import pandas as pd
from datetime import datetime
## Step 1: Save face encodings only for new names

camshots_folder = "camshots"
if not os.path.exists(camshots_folder):
    os.makedirs(camshots_folder)
    
# Initialize Face Detection and Recognition Models
print("Initializing Face Detection and Recognition Models") 
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
shape_predictor = dlib.shape_predictor("shape_predictor_68_face_landmarks.dat")
face_recognizer = dlib.face_recognition_model_v1("dlib_face_recognition_resnet_model_v1.dat")

# Step 1: Save face encodings only for new names
def load_known_faces(folder_path):
    known_face_encodings = []
    known_face_names = []

    if os.path.exists("known_faces.npy") and os.path.exists("known_names.npy"):
        known_face_encodings = list(np.load("known_faces.npy", allow_pickle=True))
        known_face_names = list(np.load("known_names.npy", allow_pickle=True))

    image_files = [f for f in os.listdir(folder_path) if f.endswith((".jpg", ".jpeg", ".png"))]

    for filename in tqdm(image_files, desc="Encoding Faces"):
        name = os.path.splitext(filename)[0]
        if name in known_face_names:
            print(f"{name} is already registered. Choose a different name.")
            continue

        image_path = os.path.join(folder_path, filename)
        image = cv2.imread(image_path)
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

        for (x, y, w, h) in faces:
            dlib_rect = dlib.rectangle(int(x), int(y), int(x + w), int(y + h))
            shape = shape_predictor(rgb_image, dlib_rect)
            face_encoding = np.array(face_recognizer.compute_face_descriptor(rgb_image, shape))

            known_face_encodings.append(face_encoding)
            known_face_names.append(name)

    np.save("known_faces.npy", np.array(known_face_encodings))
    np.save("known_names.npy", np.array(known_face_names))
    
    return known_face_encodings, known_face_names

def compare_faces(known_face_encodings, face_encoding, tolerance=0.6):
    return list(np.linalg.norm(np.array(known_face_encodings) - np.array(face_encoding), axis=1) <= tolerance)

def face_distance(known_face_encodings, face_encoding):
    return np.linalg.norm(np.array(known_face_encodings) - np.array(face_encoding), axis=1)

# Step 2: Save photo and update Excel on capture
def update_excel(name, timestamp, image_path):
    excel_file = "recognition_log.xlsx"
    
    try:
        if not os.path.exists(excel_file):
            df = pd.DataFrame(columns=["Name", "Timestamp", "Image Path"])
        else:
            df = pd.read_excel(excel_file, engine='openpyxl')
    except (ValueError, FileNotFoundError, BadZipFile):
        print(f"Error with the Excel file. Creating a new one.")
        df = pd.DataFrame(columns=["Name", "Timestamp", "Image Path"])

    # Add the new entry (name, timestamp, image path) to the DataFrame
    new_row = pd.DataFrame({
        "Name": [name], 
        "Timestamp": [timestamp], 
        "Image Path": [image_path]
    })
    
    # Append the new row to the existing DataFrame
    df = pd.concat([df, new_row], ignore_index=True)
    
    # Save the updated DataFrame to Excel
    df.to_excel(excel_file, index=False, engine='openpyxl')

    # Apply conditional formatting for red and black text
    apply_conditional_formatting(excel_file)

from openpyxl import load_workbook
from openpyxl.styles import Font

def apply_conditional_formatting(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):  # Iterate through the "Name" column
        for cell in row:
            if cell.value == "Unknown":
                cell.font = Font(color="FF0000")  # Red for "Unknown"
            else:
                cell.font = Font(color="000000")  # Black for recognized names
    
    wb.save(excel_file)
    wb.close()

key = cv2.waitKey(1)
if key == ord('s'):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    photo_filename = f"{timestamp}.jpg"
    photo_path = os.path.join(camshots_folder, photo_filename)
    
    cv2.imwrite(photo_path, frame)
    
    # Update Excel without the "Color" column
    update_excel(name, timestamp, photo_path)
    print(f"Photo saved and log updated for {name}")

def update_excel_with_image(name, timestamp, image):
    excel_file = "recognition_log.xlsx"
    
    try:
        if not os.path.exists(excel_file):
            df = pd.DataFrame(columns=["Name", "Timestamp", "Image Path"])
        else:
            df = pd.read_excel(excel_file)
    except Exception as e:
        print(f"Error with the Excel file: {e}. Creating a new one.")
        df = pd.DataFrame(columns=["Name", "Timestamp", "Image Path"])

    # Save the image in the camshots folder
    image_filename = f"{timestamp.replace(':', '-')}.jpg"
    image_path = os.path.join("camshots", image_filename)
    cv2.imwrite(image_path, image)

    # Add the new entry to the DataFrame
    new_row = pd.DataFrame({
        "Name": [name],
        "Timestamp": [timestamp],
        "Image Path": [image_path]
    })
    
    # Append the new row to the existing DataFrame
    df = pd.concat([df, new_row], ignore_index=True)
    
    # Save the updated DataFrame to Excel
    df.to_excel(excel_file, index=False)

    # Apply conditional formatting for red and black text
    apply_conditional_formatting(excel_file)

# Update the apply_conditional_formatting function to include the "Image Path" column
def apply_conditional_formatting(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        name_cell, _, path_cell = row
        if name_cell.value == "Unknown":
            name_cell.font = Font(color="FF0000")  # Red for "Unknown"
        else:
            name_cell.font = Font(color="000000")  # Black for recognized names
        
        # Make the image path a hyperlink
        path_cell.hyperlink = path_cell.value
        path_cell.style = "Hyperlink"
    
    wb.save(excel_file)
    wb.close()


def real_time_face_recognition(known_face_encodings, known_face_names):
    camshots_folder = "camshots"
    if not os.path.exists(camshots_folder):
        os.makedirs(camshots_folder)
    
    video_capture = cv2.VideoCapture(0)

    while True:
        ret, frame = video_capture.read()
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

        for (x, y, w, h) in faces:
            dlib_rect = dlib.rectangle(int(x), int(y), int(x + w), int(y + h))
            shape = shape_predictor(rgb_frame, dlib_rect)
            face_encoding = np.array(face_recognizer.compute_face_descriptor(rgb_frame, shape))

            distances = face_distance(known_face_encodings, face_encoding)
            min_distance = min(distances) if len(distances) > 0 else 1.0
            accuracy = (1 - min_distance) * 100

            name = "Unknown"
            color = (0, 0, 255)  # Red for unknown faces

            if min_distance <= 0.47:  # 53% threshold (1 - 0.53 = 0.47)
                best_match_index = np.argmin(distances)
                name = known_face_names[best_match_index]
                color = (0, 255, 0)  # Green for known faces

            # Display name and accuracy
            label = f"{name} ({accuracy:.2f}%)"
            cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
            cv2.putText(frame, label, (x + 6, y+h + 20), cv2.FONT_HERSHEY_DUPLEX, 0.8, (255, 255, 255), 1)

        cv2.imshow('Video', frame)

        key = cv2.waitKey(1)
        if key == ord('q'):
            break
        elif key == ord('s'):
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Update Excel with name, timestamp, and embedded image
            update_excel_with_image(name, timestamp, frame)
            print(f"Photo saved and log updated for {name}")

    video_capture.release()
    cv2.destroyAllWindows()


# ... (keep the rest of your code, including the main function)    
# Step 3: Delete Known Face
def delete_known_face():
    known_face_names = list(np.load("known_names.npy", allow_pickle=True))

    if not known_face_names:
        print("No known faces to delete.")
        return

    print("\nRegistered Known Faces:")
    for i, name in enumerate(known_face_names, 1):
        print(f"{i}. {name}")
    
    choice = input("\nEnter the number of the face you want to delete (or press Enter to skip): ")
    if choice.isdigit() and 1 <= int(choice) <= len(known_face_names):
        index = int(choice) - 1
        known_face_names.pop(index)
        known_face_encodings = list(np.load("known_faces.npy", allow_pickle=True))
        known_face_encodings.pop(index)
        np.save("known_names.npy", known_face_names)
        np.save("known_faces.npy", known_face_encodings)
        print("Face deleted successfully.")
    else:
        print("Skipping deletion.")
 
# Main function
def main():
    print("Face Recognition System")
    print("=======================")

    delete_choice = input("\nDo you want to delete any known face before starting? (y/n): ").lower()
    if delete_choice == 'y':
        delete_known_face()

    known_faces_folder = "Images"
    known_face_encodings, known_face_names = load_known_faces(known_faces_folder)

    print("\nKnown Faces Loaded:")
    for i, name in enumerate(known_face_names, 1):
        print(f"{i}. {name}")

    input("\nPress Enter to start real-time face recognition...")
    real_time_face_recognition(known_face_encodings, known_face_names)

if __name__ == "__main__":
    main()
